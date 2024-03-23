import base64
import json
import os
import re

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from google.oauth2 import service_account
from openpyxl import Workbook
from pydantic import BaseModel
from datetime import datetime
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side

from supabase import create_client, Client

app = FastAPI()

# Load environment variables
load_dotenv(dotenv_path=".env.local")
GOOGLE_SERVICE_KEY = os.getenv("GOOGLE_SERVICE_KEY")
if GOOGLE_SERVICE_KEY is not None:
    CREDENTIAL_STR = base64.b64decode(GOOGLE_SERVICE_KEY)
    CREDENTIAL_DICT = json.loads(CREDENTIAL_STR)
    CREDENTIALS = service_account.Credentials.from_service_account_info(CREDENTIAL_DICT)
else:
    print("GOOGLE_SERVICE_KEY could not be loaded from the environment .env.local")

url: str = os.getenv("SUPABASE_URL")
key: str = os.getenv("SUPABASE_ANON_KEY")
supabase: Client = create_client(url, key)

debug: str = os.getenv("DEBUG")
scale_ticket_table_name: str = "snapscale_scaleticket"
if debug == "TRUE":
    print("Debug mode enabled")
    scale_ticket_table_name = "snapscale_scaleticketdev"


# API Classes
class ScaleTicket:
    def __init__(
        self,
        date,
        ticket,
        gross=0.0,
        tare=0.0,
        mo=0.0,
        tw=0.0,
        driver="",
        truck="",
        sale_location="",
        field_number="",
    ):
        self.date = date
        self.ticket = ticket
        self.gross = float(gross.replace(",", "")) if isinstance(gross, str) else gross
        self.tare = float(tare.replace(",", "")) if isinstance(tare, str) else tare
        self.mo = float(mo) if isinstance(mo, str) else mo
        self.tw = float(tw) if isinstance(tw, str) else tw
        self.net = self.gross - self.tare
        self.wetBu = self.net / 56
        self.dryBu = (
            (1 - ((self.mo - 15.5) * 0.012)) * self.net / 56
            if mo > 15.5
            else self.wetBu
        )
        self.driver = driver
        self.truck = truck
        self.sale_location = sale_location
        self.field_number = field_number

    def __str__(self):
        return (
            f"ScaleTicket({self.date}, {self.sale_location}, {self.ticket}, "
            f"gross={self.gross}, tare={self.tare}, "
            f"mo={self.mo}, tw={self.tw}, "
            f"driver={self.driver}, truck={self.truck}, field_number={self.field_number})"
        )

    def __dict__(self):
        return {
            "date": self.date.strftime("%Y-%m-%d"),
            "ticket": self.ticket,
            "gross": self.gross,
            "tare": self.tare,
            "mo": self.mo,
            "tw": self.tw,
            "net": self.net,
            "wetBu": self.wetBu,
            "dryBu": self.dryBu,
            "driver": self.driver,
            "truck": self.truck,
            "sale_location": self.sale_location,
            "field_number": self.field_number,
        }

# Database functions
def upsert_scale_ticket(scale_ticket: ScaleTicket, tickets: list[ScaleTicket]):
    """
    Updates a scale ticket's database record if its ticket number and location match an existing record
    Otherwise, appends the scale ticket into the tickets list to be inserted later.
    """
    data, count = (
        supabase.table(scale_ticket_table_name)
        .select("*")
        .eq("ticket", scale_ticket.ticket)
        .execute()
    )
    print("Number of records with same ticket number", len(data[1]))
    if len(data[1]) == 0:
        # If there is no record with the same ticket number, then append the record to tickets to be inserted later.
        print("Appending new ticket number record to tickets to be inserted later")
        tickets.append(scale_ticket)
    else:
        for record in data[1]:
            print("Existing record", record)
            # If the sale location is different then append the record to tickets to be inserted later.
            if record['sale_location'] != scale_ticket.sale_location:
                print("Appending different sale location record with same ticket number to tickets to be inserted later")
                tickets.append(scale_ticket)
            else:
                # If the ticket number and sale location are the same, then update the record.
                print("Updating existing record")
                record['date'] = scale_ticket.date.strftime("%Y-%m-%d")
                record['gross'] = scale_ticket.gross
                record['tare'] = scale_ticket.tare
                record['mo'] = scale_ticket.mo
                record['tw'] = scale_ticket.tw
                record['net'] = scale_ticket.net
                record['wetBu'] = scale_ticket.wetBu
                record['dryBu'] = scale_ticket.dryBu
                record['driver'] = scale_ticket.driver
                record['truck'] = scale_ticket.truck
                record['sale_location'] = scale_ticket.sale_location
                record['field_number'] = scale_ticket.field_number
                record['status'] = "ACTIVE"
                print("Updated record", record)
                response = supabase.table(scale_ticket_table_name).update(record).eq(
                    "ticket", scale_ticket.ticket
                ).execute()
                print("Update existing record response", response)


# API functions
def perform_ocr(
    texts,
    tickets,
    location,
    ticket_regex,
    date_regex=None,
    gross_regex=None,
    tare_regex=None,
    mo_regex=None,
    tw_regex=None,
):
    """
    Perform Optical Character Recognition (OCR) on texts to extract specific information such as date,
    ticket number, gross weight, tare weight, moisture content, and total weight. It utilizes various
    regular expressions to extract the required data from the text descriptions.

    Regex patterns can be customized if needed.
    """

    date = "01/01/01"
    ticket = "000000"
    gross = "60,000"
    tare = "20,000"
    mo = 10.00
    tw = 40.00

    # Default Regular expressions
    date_regex = re.compile(r"\d{1,2}[-/\.]\d{1,2}[-/\.]\d{4}")
    gross_regex = re.compile(r"^[89]\d{1},\d{3}")
    tare_regex = re.compile(r"^[2]\d{1}\,\d{3}")
    mo_regex = re.compile(r"^[1]\d{1}\.\d{2}")
    tw_regex = re.compile(r"^[5]\d{1}\.\d{2}")

    for i, text in enumerate(texts):
        # print(f"{i}: " + text.description)
        # find date
        date_found = date_regex.search(text.description)
        if date_found:
            date_str = date_found.group()
            date = datetime.strptime(date_str, "%m/%d/%Y")

        # find ticket
        ticket_found = ticket_regex.search(text.description)
        if ticket_found:
            ticket = ticket_found.group()

        # find gross
        gross_found = gross_regex.search(text.description)
        if gross_found:
            gross = float(gross_found.group().replace(",", ""))

        # find tare
        tare_found = tare_regex.search(text.description)
        if tare_found:
            tare = float(tare_found.group().replace(",", ""))

        # find mo
        mo_found = mo_regex.search(text.description)
        if mo_found:
            mo = float(mo_found.group())

        # find tw
        tw_found = tw_regex.search(text.description)
        if tw_found:
            tw = float(tw_found.group())

    scale_ticket = ScaleTicket(
        date, ticket, gross, tare, mo, tw, sale_location=location
    )
    print(scale_ticket)
    upsert_scale_ticket(scale_ticket, tickets)


def process_images_as_scale_ticket(image_urls):
    """Detects text as a ScaleTicket in an image."""
    from google.cloud import vision_v1
    from google.cloud.vision_v1 import types

    if image_urls:
        requests = []
        tickets = []

        for image_uri in image_urls:
            image = types.Image()
            image.source.image_uri = image_uri
            requests.append(
                types.AnnotateImageRequest(
                    image=image,
                    features=[
                        types.Feature(type=vision_v1.Feature.Type.TEXT_DETECTION)
                    ],
                )
            )
        if CREDENTIALS is not None:
            client = vision_v1.ImageAnnotatorClient(credentials=CREDENTIALS)
            response = client.batch_annotate_images(requests=requests)

            if response is None:
                raise Exception(f"Batch response is None {response}")
            for i, resp in enumerate(response.responses):
                if resp.error.message:
                    print(resp.error)
                    raise Exception(
                        "{}\nFor more info on error messages, check: "
                        "https://cloud.google.com/apis/design/errors".format(
                            resp.error.message
                        )
                    )

                locations = [
                    "Frenchman Valley Coop",
                    "Baney",
                    "Imperial Beef",
                ]
                fvc_locations = [
                    "Oakley",
                    "Trenton",
                    "Imperial",
                ]
                sale_location = ""

                texts = resp.text_annotations
                for i, text in enumerate(texts):
                    sale_location = ""

                    # find location
                    for location in locations:
                        print("Checking location: " + location)
                        if location in text.description:
                            print("Found location: " + location)
                            sale_location = location
                            break
                    if sale_location != "":
                        break

                if sale_location == "Frenchman Valley Coop":
                    # find fvc location
                    for i, text in enumerate(texts):
                        sale_location = ""
                        for location in fvc_locations:
                            print("Checking fvc location: " + location)
                            if location in text.description:
                                print("Found fvc location: " + location)
                                sale_location = location
                                break
                        if sale_location != "":
                            break

                    if sale_location == "Imperial":
                        # Process as FVC Imperial
                        perform_ocr(
                            texts,
                            tickets,
                            location="FVC Imperial",
                            ticket_regex=re.compile(r"^[5][0][5]\d{3}"),
                        )
                    elif sale_location == "Oakley":
                        # Process as FVC Oakley
                        perform_ocr(
                            texts,
                            tickets,
                            location="Western Plains Energy",
                            ticket_regex=re.compile(r"^[5][3][9]\d{3}"),
                        )
                    elif sale_location == "Trenton":
                        # Process as FVC Trenton
                        perform_ocr(
                            texts,
                            tickets,
                            location="Trenton Agri Products",
                            ticket_regex=re.compile(r"^[4][0][1]\d{4}"),
                        )
                elif sale_location == "Baney":
                    # Process as Baney
                    perform_ocr(
                        texts,
                        tickets,
                        location="Baney",
                        ticket_regex=re.compile(r"^[0]\d{3}"),
                    )
                elif sale_location == "Imperial Beef":
                    # Process as Imperial Beef
                    perform_ocr(
                        texts,
                        tickets,
                        location="Imperial Beef",
                        ticket_regex=re.compile(r"^[4][0][1]\d{3}"),
                    )
                else:
                    # Process as unknown location
                    perform_ocr(
                        texts,
                        tickets,
                        location="",
                        ticket_regex=re.compile(r"\d+"),
                    )

            # TODO: save scale tickets to the supabase database
            print("Serializing tickets into json")
            dict_tickets = [t.__dict__() for t in tickets]
            print("Inserting tickets into the database: ",  dict_tickets)
            response = supabase.table(scale_ticket_table_name).insert(dict_tickets).execute()
            return dict_tickets
        else:
            raise Exception("Must provide Google Cloud credentials.")
    else:
        raise Exception("Must provide image urls to process.")


# API Excel Utils
def write_header(worksheet, location="", field=""):
    """
    Writes the header section of the workbook including contact info, location, and headers.
    Also, sets the column widths and row heights of the worksheet.
    """

    # Settings for column widths
    column_width_settings = {
        "A": 10.29 * (10.29 / 9.57),
        "B": 8 * (8 / 7.29),
        "C": 9.43 * (9.43 / 8.71),
        "D": 8 * (8 / 7.29),
        "E": 5.86 * (5.86 / 5.14),
        "F": 6.29 * (6.29 / 5.57),
        "G": 9.71 * (9.71 / 9),
        "H": 11.86 * (11.86 / 11.14),
        "I": 10.86 * (10.86 / 10.14),
        "J": 10.29 * (10.29 / 9.57),
        "K": 6.57 * (6.57 / 5.86),
    }

    # Apply column width settings
    for column, width in column_width_settings.items():
        worksheet.column_dimensions[column].width = width

    # Set row heights
    for row in range(4, 11):
        worksheet.row_dimensions[row].height = 12.75

    write_contact_info(worksheet)
    write_location_labels(worksheet)
    write_sale_location(worksheet, location)
    write_field(worksheet, field)
    write_headers(worksheet)

    return worksheet


def write_contact_info(worksheet):
    """
    Writes contact information to a worksheet in a workbook.
    """
    # Define style
    contact_style = NamedStyle(name="contact")
    contact_style.font = Font(name="Arial", size=24)
    contact_style.alignment = Alignment(horizontal="left", wrapText=False)

    # Write contact info
    contact_info = {"A1": "Jeff Wallin", "A2": "PO Box 240", "A3": "Imperial NE 69033"}

    for cell_id, info in contact_info.items():
        worksheet[cell_id] = info
        worksheet[cell_id].style = contact_style

    # Define border style
    border_side = Side(border_style="thin", color="000000")

    # Apply border style
    border_cells = {
        "A1": {"left": border_side, "top": border_side},
        "B1": {"top": border_side},
        "C1": {"right": border_side, "top": border_side},
        "A2": {"left": border_side},
        "C2": {"right": border_side},
        "A3": {"left": border_side, "bottom": border_side},
        "B3": {"bottom": border_side},
        "C3": {"right": border_side, "bottom": border_side},
    }

    for cell_id, border_sides in border_cells.items():
        worksheet[cell_id].border = Border(**border_sides)


def write_location_labels(worksheet):
    """
    Set location labels
    """
    worksheet["C6"] = "Sold To:"
    worksheet["C6"].font = Font(name="Arial", size=8)
    worksheet["C6"].alignment = Alignment(horizontal="center")

    worksheet["H6"] = "Field"
    worksheet["H6"].font = Font(name="Arial", size=10)
    worksheet["H6"].alignment = Alignment(horizontal="center")


def write_sale_location(worksheet, feedlot_name=""):
    """
    Writes the name of the sale location (feedlot name) to a specific cell in the worksheet.
    """
    worksheet["C7"] = feedlot_name
    worksheet["C7"].font = Font(name="Arial", size=11)
    worksheet["C7"].alignment = Alignment(horizontal="center")


def write_field(worksheet, field_number=""):
    """
    Writes the name of the sale location (feedlot name) to a specific cell in the worksheet.
    """
    worksheet["H7"] = field_number
    worksheet["H7"].font = Font(name="Arial", size=11)
    worksheet["H7"].alignment = Alignment(horizontal="center")


def write_headers(worksheet):
    from openpyxl.utils import get_column_letter

    # Define style
    header_style = NamedStyle(name="header")
    header_style.font = Font(name="Arial", size=10, bold=True)
    header_style.alignment = Alignment(horizontal="center")

    # Headers configuration
    headers = [
        "Date",
        "Ticket",
        "Gross",
        "Tare",
        "MO",
        "TW",
        "Net",
        "Wet Bu",
        "Dry Bu",
        "Driver",
        "Truck",
    ]

    # Implement headers
    for col_num, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_num)
        worksheet[column_letter + "10"] = header
        worksheet[column_letter + "10"].style = header_style


def write_tickets(worksheet, tickets, start_row=12):
    """Writes ticket data to the worksheet."""
    import string

    # Define Ticket style
    ticket_style = NamedStyle(name="ticket")
    ticket_style.font = Font(name="Arial", size=10)
    ticket_style.alignment = Alignment(horizontal="center")

    attributes_order = [
        "date",
        "ticket",
        "gross",
        "tare",
        "mo",
        "tw",
        "net",
        "wetBu",
        "dryBu",
        "driver",
        "truck",
    ]

    # Center Alignment for specific columns
    center_aligned_text = Alignment(horizontal="center")

    # Attribute actions for specific columns
    attribute_actions = {
        "wetBu": lambda val: round(val, 2),
        "dryBu": lambda val: round(val, 2),
        "ticket": lambda val: val,
        "gross": round,
        "tare": round,
        "net": round,
        "mo": lambda val: round(val, 1),
        "tw": lambda val: round(val, 1),
    }

    for i, ticket in enumerate(tickets, start=start_row):
        letters = iter(string.ascii_uppercase)
        for letter, attr in zip(letters, attributes_order):
            attr_value = getattr(ticket, attr)
            cell_value = attribute_actions.get(attr, lambda x: x)(attr_value)
            cell = worksheet[letter + str(i)]
            cell.value = cell_value
            cell.style = ticket_style
            if attr == "net":
                formula = f"=C{i}-D{i}"
                cell.value = formula
            elif attr == "wetBu":
                formula = f"=G{i}/56"
                cell.value = formula
            elif attr == "dryBu":
                formula = f"=IF(E{i}>15.5,(1-((E{i}-15.5)*0.012))*G{i}/56,H{i})"
                cell.value = formula
            if attr == "wetBu" or attr == "dryBu":
                cell.alignment = center_aligned_text
                cell.number_format = "#,##0.00"  # Include comma for thousands
            if attr == "date":
                cell.number_format = "mm/dd/yyyy"


def write_totals(worksheet, numTickets):
    """
    Writes the totals to the worksheet.
    """

    # Total style definition
    total_style = NamedStyle(name="total")
    total_style.font = Font(name="Arial", size=12, bold=True)
    center_aligned_text = Alignment(horizontal="center")
    total_style.alignment = center_aligned_text

    # The insertion row should be numTickets + 13 (1 rows below the last ticket).
    row = numTickets + 13

    worksheet["A" + str(row)] = "Totals"
    worksheet["A" + str(row)].style = total_style

    for letter in "CDGHI":
        start_cell = letter + "12"  # Data starts at row 12
        end_cell = letter + str(numTickets + 11)  # +11 because data starts at row 12
        formula = f"=SUM({start_cell}:{end_cell})"
        cell = worksheet[letter + str(row)]
        cell.value = formula
        cell.style = total_style
        if letter == "H" or letter == "I":
            cell.number_format = "##,##0.00"  # Include comma for thousands

    row += 2
    worksheet["A" + str(row)] = "Acres"
    worksheet["A" + str(row)].alignment = center_aligned_text
    row += 1
    worksheet["A" + str(row)] = "Yield"
    worksheet["A" + str(row)].alignment = center_aligned_text


def write_sheet(worksheet, tickets, location="", field=""):
    """
    Write the template to the given worksheet and fill it in with tickets, location, and field data.
    """
    write_header(worksheet, location, field)
    write_tickets(worksheet, tickets)
    write_totals(worksheet, len(tickets))
    worksheet.page_setup.orientation = "landscape"


# API routes
class Images(BaseModel):
    urls: list[str]


@app.post("/api/python/snapscale")
async def snapscale(images: Images):
    # Process uploaded images urls in bulk
    try:
        tickets = process_images_as_scale_ticket(images.urls)
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error processing image. {e}"
        ) from e
    return {"tickets": tickets}

    # Process the text as required and write to Excel
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    write_sheet(ws, tickets, None, None)

    # Save the workbook
    output_path = "public/output/output.xlsx"
    try:
        wb.save(output_path)
    except Exception as e:
        raise HTTPException(
            status_code=500, detail="Error generating Excel file."
        ) from e

    # Return the path to the generated Excel file
    return {"tickets": tickets}


# Error Logging
import logging
from fastapi import FastAPI, Request, status
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    exc_str = f"{exc}".replace("\n", " ").replace("   ", " ")
    logging.error(f"{request}: {exc_str}")
    content = {"status_code": 10422, "message": exc_str, "data": None}
    return JSONResponse(
        content=content, status_code=status.HTTP_422_UNPROCESSABLE_ENTITY
    )
