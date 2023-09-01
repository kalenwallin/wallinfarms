from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side

from api.classes import (generate_mock_tickets, generate_random_feedlot_name,
                         generate_random_field_number)


def write_header(worksheet, location="", field=""):
    """
    Writes the header section of the workbook including contact info, location, and headers.
    Also, sets the column widths and row heights of the worksheet.
    """

    # Settings for column widths
    column_width_settings = {
        "A": 10.29 * (10.29 / 9.57),
        "B": 8 * (8 / 7.29),
        "C": 9.43 * (9.43 / 8.71),
        "D": 8 * (8 / 7.29),
        "E": 5.86 * (5.86 / 5.14),
        "F": 6.29 * (6.29 / 5.57),
        "G": 9.71 * (9.71 / 9),
        "H": 11.86 * (11.86 / 11.14),
        "I": 10.86 * (10.86 / 10.14),
        "J": 10.29 * (10.29 / 9.57),
        "K": 6.57 * (6.57 / 5.86),
    }

    # Apply column width settings
    for column, width in column_width_settings.items():
        worksheet.column_dimensions[column].width = width

    # Set row heights
    for row in range(4, 11):
        worksheet.row_dimensions[row].height = 12.75

    write_contact_info(worksheet)
    write_location_labels(worksheet)
    write_sale_location(worksheet, location)
    write_field(worksheet, field)
    write_headers(worksheet)

    return worksheet


def write_contact_info(worksheet):
    """
    Writes contact information to a worksheet in a workbook.
    """
    # Define style
    contact_style = NamedStyle(name="contact")
    contact_style.font = Font(name="Arial", size=24)
    contact_style.alignment = Alignment(horizontal="left", wrapText=False)

    # Write contact info
    contact_info = {"A1": "Jeff Wallin", "A2": "PO Box 240", "A3": "Imperial NE 69033"}

    for cell_id, info in contact_info.items():
        worksheet[cell_id] = info
        worksheet[cell_id].style = contact_style

    # Define border style
    border_side = Side(border_style="thin", color="000000")

    # Apply border style
    border_cells = {
        "A1": {"left": border_side, "top": border_side},
        "B1": {"top": border_side},
        "C1": {"right": border_side, "top": border_side},
        "A2": {"left": border_side},
        "C2": {"right": border_side},
        "A3": {"left": border_side, "bottom": border_side},
        "B3": {"bottom": border_side},
        "C3": {"right": border_side, "bottom": border_side},
    }

    for cell_id, border_sides in border_cells.items():
        worksheet[cell_id].border = Border(**border_sides)


def write_location_labels(worksheet):
    """
    Set location labels
    """
    worksheet["C6"] = "Sold To:"
    worksheet["C6"].font = Font(name="Arial", size=8)
    worksheet["C6"].alignment = Alignment(horizontal="center")

    worksheet["H6"] = "Field"
    worksheet["H6"].font = Font(name="Arial", size=10)
    worksheet["H6"].alignment = Alignment(horizontal="center")


def write_sale_location(worksheet, feedlot_name=""):
    """
    Writes the name of the sale location (feedlot name) to a specific cell in the worksheet.
    """
    worksheet["C7"] = feedlot_name
    worksheet["C7"].font = Font(name="Arial", size=11)
    worksheet["C7"].alignment = Alignment(horizontal="center")


def write_field(worksheet, field_number=""):
    """
    Writes the name of the sale location (feedlot name) to a specific cell in the worksheet.
    """
    worksheet["H7"] = field_number
    worksheet["H7"].font = Font(name="Arial", size=11)
    worksheet["H7"].alignment = Alignment(horizontal="center")


def write_headers(worksheet):
    from openpyxl.utils import get_column_letter

    # Define style
    header_style = NamedStyle(name="header")
    header_style.font = Font(name="Arial", size=10, bold=True)
    header_style.alignment = Alignment(horizontal="center")

    # Headers configuration
    headers = [
        "Date",
        "Ticket",
        "Gross",
        "Tare",
        "MO",
        "TW",
        "Net",
        "Wet Bu",
        "Dry Bu",
        "Driver",
        "Truck",
    ]

    # Implement headers
    for col_num, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_num)
        worksheet[column_letter + "10"] = header
        worksheet[column_letter + "10"].style = header_style


def write_tickets(worksheet, tickets, start_row=12):
    """Writes ticket data to the worksheet."""
    import string

    # Define Ticket style
    ticket_style = NamedStyle(name="ticket")
    ticket_style.font = Font(name="Arial", size=10)
    ticket_style.alignment = Alignment(horizontal="center")

    attributes_order = [
        "date",
        "ticket",
        "gross",
        "tare",
        "mo",
        "tw",
        "net",
        "wetBu",
        "dryBu",
        "driver",
        "truck",
    ]

    # Center Alignment for specific columns
    center_aligned_text = Alignment(horizontal="center")

    # Attribute actions for specific columns
    attribute_actions = {
        "wetBu": lambda val: round(val, 2),
        "dryBu": lambda val: round(val, 2),
        "ticket": lambda val: val,
        "gross": round,
        "tare": round,
        "net": round,
        "mo": lambda val: round(val, 1),
        "tw": lambda val: round(val, 1),
    }

    for i, ticket in enumerate(tickets, start=start_row):
        letters = iter(string.ascii_uppercase)
        for letter, attr in zip(letters, attributes_order):
            attr_value = getattr(ticket, attr)
            cell_value = attribute_actions.get(attr, lambda x: x)(attr_value)
            cell = worksheet[letter + str(i)]
            cell.value = cell_value
            cell.style = ticket_style
            if attr == "net":
                formula = f"=C{i}-D{i}"
                cell.value = formula
            elif attr == "wetBu":
                formula = f"=G{i}/56"
                cell.value = formula
            elif attr == "dryBu":
                formula = f"=IF(E{i}>15.5,(1-((E{i}-15.5)*0.012))*G{i}/56,H{i})"
                cell.value = formula
            if attr == "wetBu" or attr == "dryBu":
                cell.alignment = center_aligned_text
                cell.number_format = "#,##0.00"  # Include comma for thousands
            if attr == "date":
                cell.number_format = "mm/dd/yyyy"


def write_totals(worksheet, numTickets):
    """
    Writes the totals to the worksheet.
    """

    # Total style definition
    total_style = NamedStyle(name="total")
    total_style.font = Font(name="Arial", size=12, bold=True)
    center_aligned_text = Alignment(horizontal="center")
    total_style.alignment = center_aligned_text

    # The insertion row should be numTickets + 13 (1 rows below the last ticket).
    row = numTickets + 13

    worksheet["A" + str(row)] = "Totals"
    worksheet["A" + str(row)].style = total_style

    for letter in "CDGHI":
        start_cell = letter + "12"  # Data starts at row 12
        end_cell = letter + str(numTickets + 11)  # +11 because data starts at row 12
        formula = f"=SUM({start_cell}:{end_cell})"
        cell = worksheet[letter + str(row)]
        cell.value = formula
        cell.style = total_style
        if letter == "H" or letter == "I":
            cell.number_format = "##,##0.00"  # Include comma for thousands

    row += 2
    worksheet["A" + str(row)] = "Acres"
    worksheet["A" + str(row)].alignment = center_aligned_text
    row += 1
    worksheet["A" + str(row)] = "Yield"
    worksheet["A" + str(row)].alignment = center_aligned_text


def write_sheet(worksheet, tickets, location="", field=""):
    """
    Write the template to the given worksheet and fill it in with tickets, location, and field data.
    """
    write_header(worksheet, location, field)
    write_tickets(worksheet, tickets)
    write_totals(worksheet, len(tickets))
    worksheet.page_setup.orientation = "landscape"


# Testing
if __name__ == "__main__":
    import sys

    if sys.argv[1] == "write_header":
        workbook = Workbook()
        worksheet = workbook.active

        # Generate mock data
        random_feedlot = generate_random_feedlot_name()
        random_field_number = generate_random_field_number()

        # Write template header with mock data
        write_header(worksheet, random_feedlot, random_field_number)

        workbook.save("output/test_write_header.xlsx")
    elif sys.argv[1] == "write_sheet":
        workbook = Workbook()
        worksheet = workbook.active

        # Generate mock data
        random_feedlot = generate_random_feedlot_name()
        random_field_number = generate_random_field_number()
        num_tickets = 10
        tickets = generate_mock_tickets(num_tickets)

        # Write mock data
        write_sheet(worksheet, tickets, random_feedlot, random_field_number)

        # Save the workbook
        workbook.save("output/test_write_sheet_2.xlsx")
    else:
        print("Error: No such function exists")
